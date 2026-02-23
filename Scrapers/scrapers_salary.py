import os
import time

import pandas as pd
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.edge.options import Options
from selenium.webdriver.edge.service import Service
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.support.ui import WebDriverWait

script_dir = os.path.dirname(os.path.abspath(__file__))
folder_path = os.path.join(os.path.dirname(script_dir), "data")
if not os.path.exists(folder_path):
    os.makedirs(folder_path)

OUTPUT = os.path.join(folder_path, "NBA_Salary.xlsx")
driver_filename = "msedgedriver.exe"

## Fonction utilitaire


def _get_driver() -> webdriver.Edge:  # Configuration du driver
    options = Options()
    options.add_argument("--start-maximized")
    options.add_argument("--disable-blink-features=AutomationControlled")
    options.add_experimental_option("excludeSwitches", ["enable-automation"])
    options.add_experimental_option("useAutomationExtension", False)
    options.add_argument("--ignore-certificate-errors")

    user_agent = "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/122.0.0.0 Safari/537.36 Edg/122.0.0.0"
    options.add_argument(f"user-agent={user_agent}")

    current_dir = os.path.dirname(os.path.abspath(__file__))
    root_dir = os.path.dirname(current_dir)
    full_driver_path = os.path.join(root_dir, driver_filename)

    service = Service(executable_path=full_driver_path)
    driver = webdriver.Edge(service=service, options=options)

    driver.execute_script(
        "Object.defineProperty(navigator, 'webdriver', {get: () => undefined})"
    )
    return driver


# Onclassifie le type de contrat pour expliquer certains "bas" salaires (rookies)
def classify_contract(
    salary: int,
) -> str:
    if salary < 2_000_000:
        return "Minimum / Two-Way"
    if salary <= 16_000_000:
        return "Rookie Scale"
    if salary > 35_000_000:
        return "Max / Supermax"
    return "Standard"


## Scraping


def scrape_salaries(max_retries: int = 3) -> pd.DataFrame:

    records: list[dict] = []
    seen: set[str] = set()

    def _read_page(driver, wait) -> list[dict]:
        new = []
        wait.until(EC.presence_of_element_located((By.CSS_SELECTOR, "table tbody tr")))
        for r in driver.find_elements(By.CSS_SELECTOR, "table tbody tr"):
            try:
                cells = r.find_elements(By.TAG_NAME, "td")
                if len(cells) < 3:
                    continue
                name = cells[1].text.strip().split("\n")[0].split("(")[0].strip()
                if not name or name in seen:
                    continue
                raw = (
                    cells[2]
                    .text.replace("$", "")
                    .replace(",", "")
                    .split("\n")[0]
                    .strip()
                )
                if raw.isdigit():
                    salary = int(raw)
                    new.append(
                        {
                            "Player": name,
                            "Salary": salary,
                            "Salary_Format": f"${salary:,}",
                            "Contract_Type": classify_contract(salary),
                        }
                    )
                    seen.add(name)
            except Exception:
                continue
        return new

    driver = _get_driver()
    wait = WebDriverWait(driver, 25)
    consecutive_errors = 0

    try:
        driver.get("https://hoopshype.com/salaries/players/")
        time.sleep(3)

        while True:
            try:
                records.extend(_read_page(driver, wait))
                consecutive_errors = 0
            except Exception as e:
                consecutive_errors += 1
                print(
                    f"[Salary] Erreur lecture ({consecutive_errors}/{max_retries}) : {e}"
                )
                if consecutive_errors >= max_retries:
                    print("[Salary] Trop d'erreurs – arrêt.")
                    break
                try:
                    driver.quit()
                except Exception:
                    pass
                time.sleep(5)
                driver = _get_driver()
                wait = WebDriverWait(driver, 25)
                driver.get("https://hoopshype.com/salaries/players/")
                time.sleep(4)
                continue

            print(f"[Salary] Joueurs collectés : {len(records)}")

            btn = None
            for selector in [
                "button.hd3Vfp__hd3Vfp._3JhbLM__3JhbLM",
                "button[aria-label='Next page']",
                "button[data-testid='pagination-next']",
                "li.next > a",
                "a[rel='next']",
            ]:
                try:
                    btn = WebDriverWait(driver, 5).until(
                        EC.element_to_be_clickable((By.CSS_SELECTOR, selector))
                    )
                    break
                except Exception:
                    continue

            if btn is None:
                print("[Salary] Fin de pagination.")
                break

            driver.execute_script("window.scrollTo(0, 1200);")
            time.sleep(1)
            driver.execute_script("arguments[0].scrollIntoView({block:'center'});", btn)
            time.sleep(0.5)
            driver.execute_script("arguments[0].click();", btn)
            time.sleep(6)

    finally:
        try:
            driver.quit()
        except Exception:
            pass

    df = (
        pd.DataFrame(records)
        .sort_values("Salary", ascending=False)
        .reset_index(drop=True)
    )
    print(f"[Salary] ✔ {len(df)} joueurs récupérés")
    return df


## Exportation et mise en page sous excel

HEADER_FILL = PatternFill("solid", start_color="1F4E79")
HEADER_FONT = Font(bold=True, color="FFFFFF", name="Arial", size=10)
ALT_FILL = PatternFill("solid", start_color="D6E4F0")
BODY_FONT = Font(name="Arial", size=10)
CENTER = Alignment(horizontal="center", vertical="center")
THIN = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)


def style_sheet(ws, df: pd.DataFrame):
    for col_idx, col_name in enumerate(df.columns, 1):
        c = ws.cell(row=1, column=col_idx, value=col_name)
        c.fill, c.font, c.alignment, c.border = HEADER_FILL, HEADER_FONT, CENTER, THIN

    for row_idx, row in enumerate(df.itertuples(index=False), 2):
        fill = ALT_FILL if row_idx % 2 == 0 else PatternFill()
        for col_idx, value in enumerate(row, 1):
            c = ws.cell(row=row_idx, column=col_idx, value=value)
            c.font, c.alignment, c.border, c.fill = BODY_FONT, CENTER, THIN, fill

    for col_idx, col_name in enumerate(df.columns, 1):
        letter = get_column_letter(col_idx)
        max_len = (
            max(
                df.iloc[:, col_idx - 1].fillna("").astype(str).map(len).max(),
                len(col_name),
            )
            + 3
        )
        ws.column_dimensions[letter].width = min(max_len, 28)

    ws.freeze_panes = "A2"


def save(df: pd.DataFrame, filename: str = OUTPUT):
    with pd.ExcelWriter(filename, engine="openpyxl") as writer:
        df.to_excel(writer, sheet_name="Salary", index=False)
        style_sheet(writer.sheets["Salary"], df)

    top = df.iloc[0]
    print(f"\n{'=' * 55}")
    print(f"✔ Fichier créé  : {filename}")
    print("  › Onglet      : Salary")
    print(f"  › Joueurs     : {len(df)}")
    print(f"  › Mieux payé  : {top['Player']} – {top['Salary_Format']}")
    print(f"{'=' * 55}\n")


## Lancement du Scrap


if __name__ == "__main__":
    df = scrape_salaries()

    if df.empty:
        print("✘ Aucun salaire récupéré. Abandon.")
    else:
        save(df)
        print("✔ Terminé !")
