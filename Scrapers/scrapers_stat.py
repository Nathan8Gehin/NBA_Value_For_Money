import os
import time

import pandas as pd
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.edge.options import Options
from selenium.webdriver.edge.service import Service
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.support.ui import WebDriverWait

driver_filename = "msedgedriver.exe"

## Fonction utilitaire


def _get_driver():  # Configuration du dirver
    options = Options()
    options.add_argument("--start-maximized")
    options.add_argument("--disable-blink-features=AutomationControlled")
    options.add_experimental_option("excludeSwitches", ["enable-automation"])
    options.add_experimental_option("useAutomationExtension", False)
    options.add_argument("--ignore-certificate-errors")

    # User-agent pour passer inaperçu
    user_agent = "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/122.0.0.0 Safari/537.36 Edg/122.0.0.0"
    options.add_argument(f"user-agent={user_agent}")

    current_dir = os.path.dirname(os.path.abspath(__file__))
    root_dir = os.path.dirname(current_dir)
    full_driver_path = os.path.join(root_dir, driver_filename)

    service = Service(executable_path=full_driver_path)
    driver = webdriver.Edge(service=service, options=options)
    driver.execute_script(
        "Object.defineProperty(navigator, 'webdriver', {get: () => undefined})"
    )
    return driver


## Fonction de Scrap


class Player:  # Définition de toute les stats
    def __init__(self, data_row):
        # On récupère toutes les colonnes exactement là où elles sont dans le HTML
        self.Player = str(data_row[1])
        self.Team = str(data_row[2])
        self.Age = int(float(data_row[3])) if data_row[3] else 0

        self.Games_Played = int(float(data_row[4])) if data_row[4] else 0
        self.Wins = int(float(data_row[5])) if data_row[5] else 0
        self.Losses = int(float(data_row[6])) if data_row[6] else 0
        self.Win_Pct = (
            round(self.Wins / self.Games_Played, 3) if self.Games_Played > 0 else 0.0
        )

        self.Minutes = float(data_row[7]) if data_row[7] else 0.0
        self.Points = float(data_row[8]) if data_row[8] else 0.0

        # Tirs
        self.Field_Goals_Made = float(data_row[9]) if data_row[9] else 0.0
        self.Field_Goals_Attempted = float(data_row[10]) if data_row[10] else 0.0
        self.FG_Percentage = float(data_row[11]) / 100 if data_row[11] else 0.0
        self.Three_PT_Made = float(data_row[12]) if data_row[12] else 0.0
        self.Three_PT_Percentage = float(data_row[14]) / 100 if data_row[14] else 0.0
        self.FT_Percentage = float(data_row[17]) / 100 if data_row[17] else 0.0

        # Défense et Impact
        self.Offensive_Rebounds = float(data_row[18]) if data_row[18] else 0.0
        self.Defensive_Rebounds = float(data_row[19]) if data_row[19] else 0.0
        self.Total_Rebounds = float(data_row[20]) if data_row[20] else 0.0
        self.Assists = float(data_row[21]) if data_row[21] else 0.0
        self.Turnovers = float(data_row[22]) if data_row[22] else 0.0
        self.Steals = float(data_row[23]) if data_row[23] else 0.0
        self.Blocks = float(data_row[24]) if data_row[24] else 0.0
        self.Personal_Fouls = float(data_row[25]) if data_row[25] else 0.0

        # Le +/- est généralement à la fin (colonne 29 ou 30)
        self.Plus_Minus = (
            float(data_row[29]) if len(data_row) > 29 and data_row[29] else 0.0
        )

    def to_dict(self):
        return self.__dict__


def scrap_nba_stat():
    driver = _get_driver()
    wait = WebDriverWait(driver, 20)
    urlpage = "https://www.nba.com/stats/players/traditional?PerMode=PerGame"
    all_player_data = []

    try:
        driver.get(urlpage)

        # Autorisation des cookies
        try:
            bypasscookie = wait.until(
                EC.element_to_be_clickable((By.ID, "onetrust-accept-btn-handler"))
            )
            bypasscookie.click()
            time.sleep(2)
        except Exception:
            print("Info: Cookies ignorés.")

        for page in range(1, 15):
            print(f"\n--- Chargement de la page {page} ---")

            try:
                wait.until(
                    EC.presence_of_element_located(
                        (
                            By.CSS_SELECTOR,
                            'table tbody tr td a[href*="/stats/player/"]',
                        )  # Attende de l'apparatition des joueurs
                    )
                )
            except Exception:
                print("⚠️ Timeout : Les vrais joueurs ne sont pas apparus.")
                break

            driver.execute_script(
                "window.scrollTo(0, 500)"
            )  # Scroll pour bien faire apparaitre toutes les statistiques
            time.sleep(3)

            # Extraction
            rows = driver.find_elements(By.CSS_SELECTOR, "table tbody tr")
            page_count = 0

            for i, r in enumerate(rows):
                try:
                    cells = r.find_elements(By.TAG_NAME, "td")
                    if len(cells) > 25:
                        data_row = []
                        for c in cells:
                            raw_text = c.get_attribute(
                                "textContent"
                            )  # On force la lecture du texte
                            clean_text = raw_text.strip() if raw_text else ""
                            data_row.append(clean_text)

                        # Debug :
                        if page == 1 and i == 0:
                            print(
                                f"🔍 PREMIER JOUEUR DÉTECTÉ : {data_row[1]} (Équipe: {data_row[2]})"
                            )

                        if len(data_row) > 1 and data_row[1] != "":
                            all_player_data.append(Player(data_row))
                            page_count += 1
                except Exception:
                    continue

            print(f"Joueurs collectés sur la page : {page_count}")

            # Pagination (Changement de page)
            try:
                next_pg = driver.find_element(
                    By.XPATH,
                    "//button[@title='Next Page'] | //button[contains(@class, 'Pagination_button') and position()=last()]",
                )

                is_disabled = next_pg.get_attribute("disabled")
                class_attr = next_pg.get_attribute("class") or ""
                if is_disabled or "disabled" in class_attr or page == 11:
                    print("🏁 Dernière page atteinte.")
                    break

                driver.execute_script("arguments[0].scrollIntoView(true);", next_pg)
                time.sleep(1)

                driver.execute_script("arguments[0].click();", next_pg)
                print(f"Passage à la page {page + 1}...")
                time.sleep(6)

            except Exception as e:
                print(
                    f"⚠️ Bouton Suivant introuvable. Fin de l'extraction. (Erreur: {e})"
                )
                break
        return all_player_data
    finally:
        driver.quit()


## Exportation et mise en page sous excel


def style_excel(path, df):
    with pd.ExcelWriter(path, engine="openpyxl") as writer:
        df.to_excel(writer, sheet_name="Stats", index=False)
        ws = writer.sheets["Stats"]

        header_fill = PatternFill("solid", start_color="1A2A4A")
        header_font = Font(bold=True, color="FFFFFF")
        alt_fill = PatternFill("solid", start_color="D6E4F0")
        border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        )

        for col_idx in range(1, len(df.columns) + 1):
            c = ws.cell(row=1, column=col_idx)
            c.fill, c.font, c.border = header_fill, header_font, border
            ws.column_dimensions[get_column_letter(col_idx)].width = 16

        for row_idx in range(2, ws.max_row + 1):
            fill = alt_fill if row_idx % 2 == 0 else PatternFill()
            for col_idx in range(1, len(df.columns) + 1):
                c = ws.cell(row=row_idx, column=col_idx)
                c.fill, c.border = fill, border
                c.alignment = Alignment(horizontal="center")

        ws.freeze_panes = "A2"


# Lancement du scrap

if __name__ == "__main__":
    data = scrap_nba_stat()

    if data:
        df = pd.DataFrame([p.to_dict() for p in data])

        # Calcul du ratio AST/TOV, et W%
        df["AST_TOV_Ratio"] = df.apply(
            lambda x: (
                round(x["Assists"] / x["Turnovers"], 2) if x["Turnovers"] > 0 else 0
            ),
            axis=1,
        )

        script_dir = os.path.dirname(os.path.abspath(__file__))
        folder_path = os.path.join(os.path.dirname(script_dir), "data")
        if not os.path.exists(folder_path):
            os.makedirs(folder_path)

        final_path = os.path.join(folder_path, "NBA_Stat.xlsx")
        style_excel(final_path, df)
        print(f"\n🎉 EXCEL SAUVEGARDÉ ! Total : {len(df)} joueurs dans {final_path}")
    else:
        print("\n❌ ÉCHEC : 0 joueurs récupérés.")
