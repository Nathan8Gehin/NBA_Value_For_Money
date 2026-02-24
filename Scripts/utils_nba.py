import pandas as pd
from openpyxl.styles import Alignment, Font, PatternFill


def compute_metrics(df):
    df = df.copy()

    # Nettoyage et conversion
    cols_stats = [
        "Points",
        "Assists",
        "Total_Rebounds",
        "Steals",
        "Blocks",
        "Turnovers",
        "Offensive_Rebounds",
        "Defensive_Rebounds",
        "Salary",
    ]
    for c in cols_stats:
        if c in df.columns:
            df[c] = pd.to_numeric(df[c], errors="coerce").fillna(0)

    # Calcul des performances
    max_vals = df[cols_stats].max().to_dict()

    def get_perf(val, col_name):
        m = max_vals.get(col_name, 1)
        return val / m if m > 0 else 0

    df["Perf_Points"] = df["Points"].apply(lambda x: get_perf(x, "Points"))
    df["Perf_Assists"] = df["Assists"].apply(lambda x: get_perf(x, "Assists"))
    df["Perf_Total_Rebounds"] = df["Total_Rebounds"].apply(
        lambda x: get_perf(x, "Total_Rebounds")
    )
    df["Perf_Steals"] = df["Steals"].apply(lambda x: get_perf(x, "Steals"))
    df["Perf_Blocks"] = df["Blocks"].apply(lambda x: get_perf(x, "Blocks"))

    m_tov = max_vals.get("Turnovers", 1)
    df["Perf_Turnovers"] = (1 - (df["Turnovers"] / m_tov)).clip(lower=0)

    # Score de performances
    # On valorise plus les contres et interceptions car plus dur à faire que de marquer 1 point
    weights = {"pts": 1.0, "ast": 1.5, "reb": 1.2, "stl": 2.5, "blk": 2.5, "tov": 1.0}
    total_w = sum(weights.values())

    df["Performance_Score"] = (
        (df["Perf_Points"] * weights["pts"])
        + (df["Perf_Assists"] * weights["ast"])
        + (df["Perf_Total_Rebounds"] * weights["reb"])
        + (df["Perf_Steals"] * weights["stl"])
        + (df["Perf_Blocks"] * weights["blk"])
        + (df["Perf_Turnovers"] * weights["tov"])
    ) / total_w

    # Calcul des impacts

    # Impact Offensif
    df["Offensive_Impact"] = (
        (df["Perf_Points"] * 0.6 + df["Perf_Assists"] * 0.4) * 100
    ).round(1)

    # Impact Défensif
    df["Defensive_Impact"] = (
        (df["Perf_Blocks"] * 0.50)
        + (df["Perf_Total_Rebounds"] * 0.35)
        + (df["Perf_Steals"] * 0.15)
    ) * 100
    df["Defensive_Impact"] = df["Defensive_Impact"].round(1)

    # Impact Global
    df["Overall_Impact"] = (df["Performance_Score"] * 100).round(1)

    # Salaire théorique
    max_perf = df["Performance_Score"].max()
    max_sal = df["Salary"].max()
    df["Salary_th"] = ((df["Performance_Score"] / max_perf) * max_sal).astype(int)
    df["Salary_th_Format"] = df["Salary_th"].apply(lambda x: f"${x:,}")
    df["Salary_Format"] = df["Salary"].apply(lambda x: f"${int(x):,}")

    # Indicateur
    def get_label(row):
        s, st = row["Salary"], row["Salary_th"]
        if st == 0 or s == 0:
            return "🟡 Well paid"
        if s > 1.25 * st:
            return "🔴 Overpaid"
        if s < 0.75 * st:
            return "🟢 Underpaid"
        return "🟡 Well paid"

    df["Indicateur"] = df.apply(get_label, axis=1)
    return df


def apply_style(ws, df):

    fill = PatternFill("solid", start_color="1F4E79")
    font = Font(bold=True, color="FFFFFF")
    for cell in ws[1]:
        cell.fill = fill
        cell.font = font
        cell.alignment = Alignment(horizontal="center")
